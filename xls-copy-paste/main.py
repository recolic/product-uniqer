
import openpyxl
import config
import os, time, logging

# https://yagisanatode.com/2017/11/18/copy-and-paste-ranges-in-excel-with-openpyxl-and-python-3/

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startRow, startCol, endRow, endCol, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startRow, startCol, endRow, endCol, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def delete_rows_below(sheet, rowIndex):
    # including rowIndex
    sheet.delete_rows(rowIndex, sheet.max_row - rowIndex + 1)


def enlarge_2darray_by_title(input_title, output_title, input_2darray):
    output_2darray = []
    
    def locate(t):
        try:
            index = output_title.index(t)
        except ValueError:
            raise RuntimeError('Unable to locate {} in output_titles {}'.format(t, output_title))
        return index
    # conv_list[index_in_input_arr] => index_in_output_arr
    conv_list = [locate(t) for t in input_title]

    def conv_1darr(arr):
        out_arr = ['' for _ in output_title]
        for cter, item in enumerate(arr):
            out_arr[conv_list[cter]] = item
        return out_arr
    return [conv_1darr(arr) for arr in input_2darray]

def get_flist(rootdir):
    result = []
    for (dirpath, dirnames, filenames) in os.walk(rootdir):
        result += [dirpath + os.path.sep + fname for fname in filenames]
    # remove the output sheet
    is_xlsx = lambda fname: fname.endswith('.xlsm') or fname.endswith('.xlsx') or fname.endswith('.xls') or fname.endswith('.XLSM') or fname.endswith('.XLSX') or fname.endswith('.XLS')
    result = list(filter(lambda f: os.path.basename(config.dst_filename) not in f and is_xlsx(f), result))
    return set(result)

import plugin_date
# (x,y) means (rowIndex, colIndex)
all_copied_data = []
def process_all(flist):
    global all_copied_data
    all_copied_data = []

    output_xls = openpyxl.load_workbook(config.template_filename)
    output_xls_sheet = output_xls.active
    output_title = copyRange(config.dst_title_ULcorner[0], config.dst_title_ULcorner[1], config.dst_title_ULcorner[0]+1, config.dst_title_ULcorner[1]+config.dst_cols, output_xls_sheet)[0]

    def process_one_src(fname):
        global all_copied_data, nt_err_msg
        logging.info('Working on file ' + fname)
        input_sheet = openpyxl.load_workbook(fname).active
        x,y = init_x,init_y = config.src_ULcorner
        while input_sheet.cell(x, y).value != None and input_sheet.cell(x, y).value != '':
            x += 1
        copied = copyRange(init_x, init_y, x, y+config.src_cols, input_sheet)

        input_title = copyRange(config.src_title_ULcorner[0], config.src_title_ULcorner[1], config.src_title_ULcorner[0]+1, config.src_title_ULcorner[1]+config.src_cols, input_sheet)[0]
        copied, input_title = plugin_date.add_date_col(input_sheet.cell(1,1).value, copied, input_title) # dirty: generate date col.
        all_copied_data += enlarge_2darray_by_title(input_title, output_title, copied)

        # silly error report
        if list(filter(lambda ls: None in ls, copied)) != []:
            # contains None
            nt_err_msg += 'Found `None` in file ' + fname + '\r\n'

    # iterate over files
    sorted_flist = list(flist)
    sorted_flist.sort()
    for fname in sorted_flist:
        process_one_src(fname)

    logging.info(all_copied_data)
    if len(all_copied_data) > 0:
        output_x, output_y = config.dst_ULcorner # starts from 1
        pasteRange(output_x, output_y, output_x + len(all_copied_data), output_y + len(all_copied_data[0]), output_xls_sheet, all_copied_data)

    delete_rows_below(output_xls_sheet, config.dst_ULcorner[0] + len(all_copied_data))
    output_xls.save(config.dst_filename)

logging.basicConfig(format='[%(asctime)s] %(message)s', level=logging.INFO)
nt_err_msg = ''
def daemon_main(argv):
    logging.info('Daemon running...')
    prev_flist = set()
    while True:
        flist = get_flist(config.src_dir_path)
        if prev_flist != flist:
            logging.info('Detected file change. Re-calculating...')
            process_all(flist)
        prev_flist = flist

        time.sleep(config.daemon_loop_interval)

def main(argv):
    process_all(get_flist(config.src_dir_path))
    if os.name == 'nt':
        os.startfile(config.dst_filename)
        if nt_err_msg != '':
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tf:
                tf.write(nt_err_msg.encode('gb2312'))
            os.startfile(tf.name)






